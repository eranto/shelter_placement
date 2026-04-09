"""
Generate an updated budget Excel for road shelters with capacity-based costing.

Cost model (unit cost per shelter, editable in dashboard):
  6 people  → 73,000 ₪
  12 people → 100,000 ₪
  20 people → 150,000 ₪
  30 people → 200,000 ₪
  50 people → 250,000 ₪

Costs reflect standard Israeli roadside shelter (מיגונית) construction,
scaled by capacity. All costs excl. VAT.

Input:  shelters_final_placements.csv  (already border-filtered and capacity-upgraded)
Output: budget_shelters_v2.xlsx
"""

import csv, math
from pathlib import Path
from collections import Counter, defaultdict

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("pip install openpyxl")

HERE = Path("/Users/erantoch/My Drive (erantoch@gmail.com)/Public Work/code/Shelter Placement 2006")

# ── Cost model (user-set prices) ─────────────────────────────────────────────
CAPACITY_COST = {6: 73_000, 12: 100_000, 20: 150_000}
MAINTENANCE_ANNUAL = 15_000   # ₪ per shelter per year
RESERVE_PCT = 15              # contingency %

# ── Geographic region from lat/lon ────────────────────────────────────────────
# zone values in shelters_final_placements.csv are already translated:
# 'North border', 'Gaza envelope', 'Standard'
def geo_region(lat, lon, zone):
    if zone == 'North border' or lat > 33.0:
        return 'צפון (גליל / גבול לבנון)'
    if zone == 'Gaza envelope' or (lat < 31.60 and lon < 34.90):
        return 'עוטף עזה / הנגב המערבי'
    if lat > 32.5:
        return 'חיפה / כרמל'
    if lat > 32.0:
        return 'מרכז (שרון / גוש דן)'
    if lat > 31.5:
        return 'ירושלים / שפלה'
    return 'נגב / ערבה'

# ── Load shelters ─────────────────────────────────────────────────────────────
# shelters_final_placements.csv is already border-filtered and capacity-upgraded
shelters = []
with open(HERE / "shelters_final_placements.csv", newline='', encoding='utf-8') as f:
    for r in csv.DictReader(f):
        lat  = float(r['lat'])
        lon  = float(r['lon'])
        cap  = int(r['suggested_capacity'])   # already upgraded
        alerts = int(r['nearby_alerts_marapr2026']) if r.get('nearby_alerts_marapr2026') else 0
        shelters.append({
            'rank':     int(r['rank']),
            'lat':      lat,
            'lon':      lon,
            'road':     r['road'],
            'zone':     r['zone'],
            'region':   geo_region(lat, lon, r['zone']),
            'capacity': cap,
            'unit_cost': CAPACITY_COST.get(cap, 650_000),
            'aadt_normal':  int(r['aadt_normal']),
            'aadt_wartime': int(r['aadt_wartime']),
            'people':   float(r['estimated_people_in_catchment']),
            'speed':    int(r['road_speed_kmh']),
            'alerts':   alerts,
        })

print(f"Loaded {len(shelters)} shelters.")

# ── Helpers ───────────────────────────────────────────────────────────────────
def hdr_fill(color='1A535C'):
    return PatternFill("solid", fgColor=color)

def yellow_fill():
    return PatternFill("solid", fgColor='FFF3CD')

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def bold(size=11):
    return Font(bold=True, size=size)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

ILS = '#,##0 ₪'
NUM = '#,##0'

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Dashboard
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'לוח בקרה'
ws.sheet_view.rightToLeft = True

cap_sizes  = [6, 12, 20]
cap_counts = Counter(s['capacity'] for s in shelters)
total      = len(shelters)

# Title
ws.merge_cells('A1:G1')
ws['A1'] = 'לוח בקרה — תקציב מיגוניות כבישים | ישראל'
ws['A1'].font = Font(bold=True, size=15)
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:G2')
ws['A2'] = 'שנה את התאים המסומנים בצהוב כדי לעדכן את כל החישובים'
ws['A2'].font = Font(italic=True, color='888888', size=10)
ws['A2'].alignment = Alignment(horizontal='center')

# ── Section: capacity tiers ───────────────────────────────────────────────────
row = 4
ws.cell(row, 1, 'עלות יחידה לפי קיבולת (ניתן לשינוי)').font = bold(12)
row += 1
for c, label in zip('ABCDE', ['קיבולת (אנשים)', 'עלות ליחידה (₪)', 'כמות', 'עלות כוללת (₪)', 'שורה בנוסחה']):
    ws.cell(row, 'ABCDE'.index(c)+1, label).font = bold()

row += 1
cost_rows = {}
for cap in cap_sizes:
    cnt = cap_counts.get(cap, 0)
    ws.cell(row, 1, cap)
    cost_cell = ws.cell(row, 2, CAPACITY_COST.get(cap, 0))
    cost_cell.fill = yellow_fill()
    cost_cell.number_format = ILS
    ws.cell(row, 3, cnt)
    total_cell = ws.cell(row, 4)
    total_cell.value = f'=B{row}*C{row}'
    total_cell.number_format = ILS
    cost_rows[cap] = row
    row += 1

sum_formula = '+'.join(f'D{cost_rows[cap]}' for cap in cap_sizes)
ws.cell(row, 1, 'עלות הון סה"כ').font = bold()
total_cap_row = row
ws.cell(row, 4, f'={sum_formula}').number_format = ILS
ws.cell(row, 4).font = bold()

# ── Section: parameters ───────────────────────────────────────────────────────
row += 2
ws.cell(row, 1, 'פרמטרים נוספים (ניתן לשינוי)').font = bold(12)
row += 1
ws.cell(row, 1, 'רזרבה / אי-וודאות (%)')
reserve_cell = ws.cell(row, 2, RESERVE_PCT)
reserve_cell.fill = yellow_fill()
reserve_row = row

row += 1
ws.cell(row, 1, 'עלות תחזוקה שנתית לכל מיגונית (₪)')
maint_cell = ws.cell(row, 2, MAINTENANCE_ANNUAL)
maint_cell.fill = yellow_fill()
maint_cell.number_format = ILS
maint_row = row

# ── Section: budget summary ───────────────────────────────────────────────────
row += 2
ws.cell(row, 1, 'סיכום תקציבי').font = bold(12)
row += 1
ws.cell(row, 1, 'מספר מיגוניות')
ws.cell(row, 2, total)

row += 1
ws.cell(row, 1, 'עלות הון (ללא מע"מ)')
ws.cell(row, 2, f'=D{total_cap_row}').number_format = ILS
capex_row = row

row += 1
ws.cell(row, 1, 'רזרבה')
ws.cell(row, 2, f'=D{total_cap_row}*B{reserve_row}/100').number_format = ILS

row += 1
ws.cell(row, 1, 'עלות הון כולל רזרבה')
ws.cell(row, 2, f'=D{total_cap_row}*(1+B{reserve_row}/100)').number_format = ILS
ws.cell(row, 2).font = bold()

row += 1
ws.cell(row, 1, f'עלות תחזוקה שנתית ({total} מיגוניות)')
ws.cell(row, 2, f'=B{maint_row}*{total}').number_format = ILS

row += 1
ws.cell(row, 1, 'עלות כוללת ל-10 שנים')
ws.cell(row, 2, f'=D{total_cap_row}*(1+B{reserve_row}/100)+B{maint_row}*{total}*10').number_format = ILS
ws.cell(row, 2).font = bold()

row += 2
ws.cell(row, 1, (
    'הערות: עלויות ללא מע"מ. מבוסס על אלגוריתם Gonzalez k-center | '
    f'תקן 5 דקות | {total} מיגוניות | ישראל בלבד (קו ירוק 1967) | '
    'קיבולת מחושבת לפי תנועה בשעת שיא בזמן מלחמה (10% מתנועה רגילה)'
)).font = Font(italic=True, color='888888', size=9)

for c in [1, 2, 3, 4]:
    set_col_width(ws, c, 36 if c == 1 else 20)

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Shelter list
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('מיגוניות')
ws2.sheet_view.rightToLeft = True

headers = ['#', 'קו רוחב', 'קו אורך', 'כביש', 'אזור', 'אזור גיאוגרפי',
           'קיבולת (אנשים)', 'עלות יחידה (₪)',
           'תנועה יומית רגילה', 'תנועה בזמן מלחמה',
           'מהירות (קמ"ש)', 'אנשים צפויים', 'התרעות מר-אפר 2026']

ws2.merge_cells(f'A1:{get_column_letter(len(headers))}1')
ws2['A1'] = f'רשימת מיגוניות מלאה — {total} יחידות'
ws2['A1'].font = Font(bold=True, size=13)

for col, h in enumerate(headers, 1):
    c = ws2.cell(2, col, h)
    c.font = Font(bold=True, color='FFFFFF', size=10)
    c.fill = hdr_fill()
    c.alignment = Alignment(horizontal='center')

ZONE_HE = {'North border': 'צפון', 'Gaza envelope': 'עוטף עזה', 'Standard': 'רגיל'}

for i, s in enumerate(shelters, 3):
    ws2.cell(i, 1,  s['rank'])
    ws2.cell(i, 2,  s['lat']).number_format  = '0.00000'
    ws2.cell(i, 3,  s['lon']).number_format  = '0.00000'
    ws2.cell(i, 4,  s['road'])
    ws2.cell(i, 5,  ZONE_HE.get(s['zone'], s['zone']))
    ws2.cell(i, 6,  s['region'])
    ws2.cell(i, 7,  s['capacity'])
    c = ws2.cell(i, 8,  s['unit_cost'])
    c.number_format = ILS
    ws2.cell(i, 9,  s['aadt_normal']).number_format  = NUM
    ws2.cell(i, 10, s['aadt_wartime']).number_format = NUM
    ws2.cell(i, 11, s['speed'])
    ws2.cell(i, 12, round(s['people'], 1))
    ws2.cell(i, 13, s['alerts'] if s['alerts'] else '')

    if i % 2 == 0:
        for col in range(1, len(headers)+1):
            ws2.cell(i, col).fill = PatternFill("solid", fgColor='F8F9FA')

col_widths = [5, 10, 10, 24, 10, 24, 14, 18, 18, 18, 12, 14, 20]
for col, w in enumerate(col_widths, 1):
    set_col_width(ws2, col, w)

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — By region
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('לפי אזור')
ws3.sheet_view.rightToLeft = True

region_data = defaultdict(lambda: {'count': 0, 'cost': 0})
for s in shelters:
    region_data[s['region']]['count'] += 1
    region_data[s['region']]['cost']  += s['unit_cost']

ws3['A1'] = 'פירוט לפי אזור גיאוגרפי'
ws3['A1'].font = bold(13)

headers3 = ['אזור', 'כמות מיגוניות', 'עלות כוללת (₪)', '% מסה"כ כמות', '% מסה"כ עלות']
for col, h in enumerate(headers3, 1):
    c = ws3.cell(2, col, h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = hdr_fill()

total_cost = sum(s['unit_cost'] for s in shelters)
for i, (region, d) in enumerate(sorted(region_data.items()), 3):
    ws3.cell(i, 1, region)
    ws3.cell(i, 2, d['count'])
    ws3.cell(i, 3, d['cost']).number_format = ILS
    ws3.cell(i, 4, round(d['count']/total*100, 1))
    ws3.cell(i, 5, round(d['cost']/total_cost*100, 1))

tot_row = 3 + len(region_data)
ws3.cell(tot_row, 1, 'סה"כ').font = bold()
ws3.cell(tot_row, 2, total).font = bold()
ws3.cell(tot_row, 3, total_cost).number_format = ILS
ws3.cell(tot_row, 3).font = bold()

for col, w in zip(range(1,6), [28, 14, 18, 16, 16]):
    set_col_width(ws3, col, w)

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Capacity breakdown
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('לפי קיבולת')
ws4.sheet_view.rightToLeft = True

ws4['A1'] = 'פירוט לפי קיבולת'
ws4['A1'].font = bold(13)

headers4 = ['קיבולת (אנשים)', 'כמות מיגוניות', 'עלות ליחידה (₪)', 'עלות כוללת (₪)', '% מסה"כ כמות']
for col, h in enumerate(headers4, 1):
    c = ws4.cell(2, col, h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = hdr_fill()

for i, cap in enumerate(cap_sizes, 3):
    cnt  = cap_counts.get(cap, 0)
    cost = CAPACITY_COST.get(cap, 0)
    ws4.cell(i, 1, cap)
    ws4.cell(i, 2, cnt)
    ws4.cell(i, 3, cost).number_format = ILS
    ws4.cell(i, 4, cnt * cost).number_format = ILS
    ws4.cell(i, 5, round(cnt/total*100, 1))

tot_row4 = 3 + len(cap_sizes)
ws4.cell(tot_row4, 1, 'סה"כ').font = bold()
ws4.cell(tot_row4, 2, total).font = bold()
ws4.cell(tot_row4, 4, total_cost).number_format = ILS
ws4.cell(tot_row4, 4).font = bold()

for col, w in zip(range(1,6), [16, 14, 18, 18, 16]):
    set_col_width(ws4, col, w)

# ── Save ──────────────────────────────────────────────────────────────────────
out = HERE / "budget_shelters_v2.xlsx"
wb.save(out)
print(f"Saved: {out.name}")
print(f"\nSummary:")
print(f"  Shelters: {total}")
print(f"  Total capital cost: ₪{total_cost:,}")
print(f"  With {RESERVE_PCT}% reserve: ₪{int(total_cost*1.15):,}")
print(f"  10-year total: ₪{int(total_cost*1.15 + MAINTENANCE_ANNUAL*total*10):,}")
print(f"\nCapacity breakdown:")
for cap in cap_sizes:
    cnt = cap_counts.get(cap, 0)
    print(f"  {cap:>3} people: {cnt:>3} shelters  @ ₪{CAPACITY_COST[cap]:>9,} = ₪{cnt*CAPACITY_COST[cap]:>12,}")
