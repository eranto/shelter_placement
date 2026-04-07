"""
Generates budget_shelters.xlsx — a Hebrew Excel workbook for managing
and planning road shelter (מיגוניות) budgets.

Sheets:
  1. לוח בקרה   — dashboard with editable inputs and live totals
  2. מיגוניות    — full shelter list with per-unit costs
  3. לפי אזור   — regional breakdown
  4. רגישות      — sensitivity table (what-if on unit costs)

Run:
  python create_budget_excel.py
Output: budget_shelters.xlsx (next to this script)
"""

import math
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── Load shelter data ─────────────────────────────────────────────────────────
BASE = Path(__file__).parent
print("טוען נתוני מיגוניות …")
ns = {"__file__": str(BASE / "optimize_shelters.py")}
exec(open(BASE / "optimize_shelters.py").read(), ns)
shelter_points = ns["shelter_points"]   # (lat, lon, risk)
print(f"נטענו {len(shelter_points)} מיגוניות.\n")

def speed_kmh(risk):
    if risk >= 4.0: return 70
    if risk >= 3.0: return 80
    if risk >= 2.0: return 90
    return 110

def risk_label(risk):
    if risk >= 4.0: return 'גבול / עוטף עזה'
    if risk >= 3.0: return 'סיכון גבוה'
    if risk >= 2.0: return 'כביש ראשי'
    return 'כביש מהיר'

def region_label(lat, lon):
    if lat >= 32.8:                              return 'צפון (גליל / גבול לבנון)'
    if 32.5 <= lat < 32.8:                       return 'חיפה / כרמל'
    if 32.0 <= lat < 32.5:                       return 'שרון / מרכז'
    if 31.9 <= lat < 32.1 and lon < 35.0:        return 'מטרופולין תל אביב'
    if 31.72 <= lat < 31.92 and 34.92 <= lon <= 35.25: return 'ירושלים'
    if 31.2 <= lat <= 31.7 and lon < 34.6:       return 'עוטף עזה'
    if 29.5 <= lat < 31.3:                       return 'נגב / ערבה'
    return 'אחר'

# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BLUE  = "1A237E"
MID_BLUE   = "283593"
LIGHT_BLUE = "E8EAF6"
RED        = "B71C1C"
ORANGE     = "E65100"
YELLOW_BG  = "FFF9C4"
GREEN      = "1B5E20"
WHITE      = "FFFFFF"
GREY_BG    = "F5F5F5"
BORDER_CLR = "BDBDBD"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold(size=11, color="000000", italic=False):
    return Font(name="Arial", bold=True, size=size, color=color, italic=italic)

def normal(size=10, color="000000"):
    return Font(name="Arial", size=size, color=color)

def border(style="thin"):
    s = Side(style=style, color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

def thick_border():
    s = Side(style="medium", color="555555")
    return Border(left=s, right=s, top=s, bottom=s)

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — לוח בקרה (Dashboard)
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "לוח בקרה"
ws1.sheet_view.rightToLeft = True
ws1.sheet_view.tabSelected = True

# Column widths
for col, w in [('A',28),('B',18),('C',18),('D',18),('E',18),('F',18),('G',18)]:
    ws1.column_dimensions[col].width = w
ws1.row_dimensions[1].height = 40
ws1.row_dimensions[2].height = 20

# ── Title ─────────────────────────────────────────────────────────────────────
ws1.merge_cells("A1:G1")
c = ws1["A1"]
c.value = "לוח בקרה — תקציב מיגוניות כבישים | ישראל"
c.font  = bold(18, WHITE)
c.fill  = fill(DARK_BLUE)
c.alignment = center()

ws1.merge_cells("A2:G2")
c = ws1["A2"]
c.value = "שנה את התאים המסומנים בצהוב כדי לעדכן את כל החישובים"
c.font  = normal(10, "555555")
c.fill  = fill(LIGHT_BLUE)
c.alignment = center()

# ── Section: counts by risk tier ─────────────────────────────────────────────
by_risk = defaultdict(int)
for _, _, r in shelter_points:
    by_risk[r] += 1

row = 4
ws1.merge_cells(f"A{row}:G{row}")
c = ws1.cell(row, 1, "כמות מיגוניות לפי דרגת סיכון")
c.font = bold(12, WHITE); c.fill = fill(MID_BLUE); c.alignment = center()

row += 1
headers = ["דרגת סיכון", "מהירות (קמ\"ש)", "מרווח מקסימלי (ק\"מ)", "כמות מיגוניות"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row, col, h)
    c.font = bold(10, WHITE); c.fill = fill(MID_BLUE)
    c.alignment = center(); c.border = border()

TIER_ROWS = {}
tier_colors = {4.0: "FFCDD2", 2.5: "FFE0B2", 2.0: "FFF9C4", 1.8: "DCEDC8", 1.5: "DCEDC8"}
row += 1
for risk in sorted(by_risk.keys(), reverse=True):
    spd = speed_kmh(risk)
    gap = round(2 * (5/60) * spd, 1)
    bg  = tier_colors.get(risk, WHITE)
    vals = [risk_label(risk), spd, gap, by_risk[risk]]
    for col, v in enumerate(vals, 1):
        c = ws1.cell(row, col, v)
        c.font = normal(10); c.fill = fill(bg)
        c.alignment = center(); c.border = border()
    TIER_ROWS[risk] = row
    row += 1

# Total count row
total_shelters = len(shelter_points)
ws1.cell(row, 1, "סה\"כ").font = bold()
ws1.cell(row, 1).alignment = center()
ws1.cell(row, 4, total_shelters).font = bold()
ws1.cell(row, 4).alignment = center()
for col in range(1, 5):
    ws1.cell(row, col).border = border()
    ws1.cell(row, col).fill = fill(LIGHT_BLUE)
COUNT_TOTAL_ROW = row

# ── Section: unit cost inputs ─────────────────────────────────────────────────
row += 2
ws1.merge_cells(f"A{row}:G{row}")
c = ws1.cell(row, 1, "עלות יחידה לפי דרגת סיכון (ניתן לשינוי)")
c.font = bold(12, WHITE); c.fill = fill(MID_BLUE); c.alignment = center()

row += 1
for col, h in enumerate(["דרגת סיכון", "עלות ליחידה (₪)", "כמות", "עלות כוללת (₪)"], 1):
    c = ws1.cell(row, col, h)
    c.font = bold(10, WHITE); c.fill = fill(MID_BLUE)
    c.alignment = center(); c.border = border()

default_costs = {4.0: 520_000, 2.5: 480_000, 2.0: 430_000, 1.8: 400_000, 1.5: 400_000}
COST_INPUT_ROWS = {}
row += 1
subtotal_rows = []
for risk in sorted(by_risk.keys(), reverse=True):
    bg = tier_colors.get(risk, WHITE)
    ws1.cell(row, 1, risk_label(risk)).font = normal(10)
    ws1.cell(row, 1).alignment = right(); ws1.cell(row, 1).border = border()
    ws1.cell(row, 1).fill = fill(bg)

    # Editable unit cost cell (yellow)
    uc = ws1.cell(row, 2, default_costs.get(risk, 400_000))
    uc.font = bold(10); uc.fill = fill(YELLOW_BG)
    uc.number_format = '#,##0 ₪'; uc.alignment = center(); uc.border = thick_border()

    # Count (linked to tier counts above)
    cnt = ws1.cell(row, 3, by_risk[risk])
    cnt.font = normal(10); cnt.alignment = center(); cnt.border = border()
    cnt.fill = fill(bg)

    # Subtotal formula
    sub = ws1.cell(row, 4)
    sub.value = f"=B{row}*C{row}"
    sub.number_format = '#,##0 ₪'; sub.alignment = center(); sub.border = border()
    sub.font = normal(10); sub.fill = fill(bg)
    subtotal_rows.append(row)
    COST_INPUT_ROWS[risk] = row
    row += 1

# Capital total
ws1.cell(row, 1, "עלות הון סה\"כ").font = bold()
ws1.cell(row, 1).fill = fill(LIGHT_BLUE); ws1.cell(row, 1).border = border()
cap_formula = "+".join([f"D{r}" for r in subtotal_rows])
cap_cell = ws1.cell(row, 4, f"={cap_formula}")
cap_cell.number_format = '#,##0 ₪'; cap_cell.font = bold()
cap_cell.fill = fill(LIGHT_BLUE); cap_cell.border = border(); cap_cell.alignment = center()
CAP_ROW = row

# ── Section: contingency + maintenance ───────────────────────────────────────
row += 2
ws1.merge_cells(f"A{row}:G{row}")
c = ws1.cell(row, 1, "פרמטרים נוספים (ניתן לשינוי)")
c.font = bold(12, WHITE); c.fill = fill(MID_BLUE); c.alignment = center()
row += 1

params = [
    ("רזרבה / אי-וודאות (%)", 15, "CONTINGENCY"),
    ("עלות תחזוקה שנתית לכל מיגונית (₪)", 15_000, "MAINT_UNIT"),
]
PARAM_ROWS = {}
for label, default, key in params:
    ws1.cell(row, 1, label).font = normal(10)
    ws1.cell(row, 1).alignment = right()
    ws1.cell(row, 1).border = border()
    inp = ws1.cell(row, 2, default)
    inp.font = bold(10); inp.fill = fill(YELLOW_BG)
    inp.alignment = center(); inp.border = thick_border()
    if key == "CONTINGENCY":
        inp.number_format = '0"%"'
    else:
        inp.number_format = '#,##0 ₪'
    PARAM_ROWS[key] = row
    row += 1

# ── Section: summary totals ───────────────────────────────────────────────────
row += 1
ws1.merge_cells(f"A{row}:G{row}")
c = ws1.cell(row, 1, "סיכום תקציבי")
c.font = bold(12, WHITE); c.fill = fill(DARK_BLUE); c.alignment = center()
row += 1

summary_items = [
    ("עלות הון (ללא מע\"מ)",          f"=D{CAP_ROW}"),
    ("רזרבה",                          f"=D{CAP_ROW}*B{PARAM_ROWS['CONTINGENCY']}/100"),
    ("עלות הון כולל רזרבה",            f"=D{CAP_ROW}*(1+B{PARAM_ROWS['CONTINGENCY']}/100)"),
    ("עלות תחזוקה שנתית (כל המיגוניות)", f"=B{PARAM_ROWS['MAINT_UNIT']}*{total_shelters}"),
    ("עלות כוללת ל-10 שנים",           f"=D{CAP_ROW}*(1+B{PARAM_ROWS['CONTINGENCY']}/100)+B{PARAM_ROWS['MAINT_UNIT']}*{total_shelters}*10"),
]
SUMMARY_ROWS = {}
for label, formula in summary_items:
    ws1.cell(row, 1, label).font = bold(10)
    ws1.cell(row, 1).fill = fill(LIGHT_BLUE); ws1.cell(row, 1).border = border()
    ws1.cell(row, 1).alignment = right()
    c = ws1.cell(row, 2, formula)
    c.number_format = '#,##0 ₪'; c.font = bold(11, DARK_BLUE)
    c.fill = fill(LIGHT_BLUE); c.border = thick_border(); c.alignment = center()
    SUMMARY_ROWS[label] = row
    row += 1

# ── Notes ────────────────────────────────────────────────────────────────────
row += 1
ws1.merge_cells(f"A{row}:G{row}")
c = ws1.cell(row, 1, "הערות: עלויות ללא מע\"מ. חישוב בוצע על בסיס אלגוריתם Gonzalez k-center | תקן 5 דקות | ישראל בלבד (קו ירוק 1967)")
c.font = normal(9, "777777"); c.alignment = right()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — מיגוניות (Full shelter list)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("מיגוניות")
ws2.sheet_view.rightToLeft = True

for col, w in [('A',6),('B',14),('C',14),('D',22),('E',16),('F',14),('G',22),('H',20)]:
    ws2.column_dimensions[col].width = w

ws2.merge_cells("A1:H1")
c = ws2["A1"]
c.value = f"רשימת מיגוניות מלאה — {total_shelters} יחידות"
c.font = bold(14, WHITE); c.fill = fill(DARK_BLUE); c.alignment = center()
ws2.row_dimensions[1].height = 30

headers2 = ["#", "קו רוחב", "קו אורך", "דרגת סיכון", "מהירות (קמ\"ש)", "עלות יחידה (₪)", "אזור", "הערות"]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(2, col, h)
    c.font = bold(10, WHITE); c.fill = fill(MID_BLUE)
    c.alignment = center(); c.border = border()
ws2.row_dimensions[2].height = 22

for i, (lat, lon, risk) in enumerate(shelter_points, 1):
    r = i + 2
    bg = tier_colors.get(risk, WHITE) if i % 2 == 0 else WHITE
    vals = [i, round(lat, 5), round(lon, 5), risk_label(risk),
            speed_kmh(risk), default_costs.get(risk, 400_000),
            region_label(lat, lon), ""]
    for col, v in enumerate(vals, 1):
        c = ws2.cell(r, col, v)
        c.font = normal(9); c.fill = fill(bg if i % 2 == 0 else "FAFAFA")
        c.alignment = center(); c.border = border()
        if col == 6:
            c.number_format = '#,##0 ₪'

ws2.freeze_panes = "A3"
ws2.auto_filter.ref = f"A2:H{total_shelters + 2}"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — לפי אזור (Regional breakdown)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("לפי אזור")
ws3.sheet_view.rightToLeft = True

for col, w in [('A',30),('B',16),('C',20),('D',22),('E',22)]:
    ws3.column_dimensions[col].width = w

ws3.merge_cells("A1:E1")
c = ws3["A1"]
c.value = "פירוט לפי אזור גיאוגרפי"
c.font = bold(14, WHITE); c.fill = fill(DARK_BLUE); c.alignment = center()
ws3.row_dimensions[1].height = 30

for col, h in enumerate(["אזור", "כמות מיגוניות", "עלות כוללת (₪)", "% מסה\"כ כמות", "% מסה\"כ עלות"], 1):
    c = ws3.cell(2, col, h)
    c.font = bold(10, WHITE); c.fill = fill(MID_BLUE)
    c.alignment = center(); c.border = border()

regions_data = defaultdict(lambda: {"count": 0, "cost": 0})
for lat, lon, risk in shelter_points:
    reg = region_label(lat, lon)
    regions_data[reg]["count"] += 1
    regions_data[reg]["cost"]  += default_costs.get(risk, 400_000)

total_cost = sum(v["cost"] for v in regions_data.values())

region_order = [
    'צפון (גליל / גבול לבנון)', 'חיפה / כרמל', 'שרון / מרכז',
    'מטרופולין תל אביב', 'ירושלים', 'עוטף עזה', 'נגב / ערבה', 'אחר'
]
reg_colors = ["E3F2FD","E8F5E9","FFF9C4","FCE4EC","EDE7F6","FBE9E7","F3E5F5","FAFAFA"]

for i, reg in enumerate(region_order):
    if reg not in regions_data:
        continue
    r = i + 3
    d = regions_data[reg]
    bg = reg_colors[i % len(reg_colors)]
    vals = [reg, d["count"], d["cost"],
            round(d["count"] / total_shelters * 100, 1),
            round(d["cost"]  / total_cost   * 100, 1)]
    for col, v in enumerate(vals, 1):
        c = ws3.cell(r, col, v)
        c.font = normal(10); c.fill = fill(bg)
        c.alignment = center(); c.border = border()
        if col == 3: c.number_format = '#,##0 ₪'
        if col in (4, 5): c.number_format = '0.0"%"'

last_r = len(region_order) + 3
ws3.cell(last_r, 1, "סה\"כ").font = bold()
ws3.cell(last_r, 1).fill = fill(LIGHT_BLUE); ws3.cell(last_r, 1).border = border()
ws3.cell(last_r, 2, total_shelters).font = bold()
ws3.cell(last_r, 2).number_format = '#,##0'
ws3.cell(last_r, 2).fill = fill(LIGHT_BLUE); ws3.cell(last_r, 2).border = border(); ws3.cell(last_r, 2).alignment = center()
ws3.cell(last_r, 3, total_cost).font = bold()
ws3.cell(last_r, 3).number_format = '#,##0 ₪'
ws3.cell(last_r, 3).fill = fill(LIGHT_BLUE); ws3.cell(last_r, 3).border = border(); ws3.cell(last_r, 3).alignment = center()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — רגישות (Sensitivity / what-if table)
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("רגישות")
ws4.sheet_view.rightToLeft = True

for col, w in [('A',30),('B',18),('C',18),('D',18),('E',18),('F',18)]:
    ws4.column_dimensions[col].width = w

ws4.merge_cells("A1:F1")
c = ws4["A1"]
c.value = "ניתוח רגישות — השפעת שינוי עלות יחידה על התקציב הכולל"
c.font = bold(14, WHITE); c.fill = fill(DARK_BLUE); c.alignment = center()
ws4.row_dimensions[1].height = 30

ws4.merge_cells("A2:F2")
c = ws4["A2"]
c.value = "כל שינוי בלוח הבקרה יתעדכן כאן אוטומטית"
c.font = normal(10, "555555"); c.fill = fill(LIGHT_BLUE); c.alignment = center()

# What-if table: vary unit cost multiplier from 80% to 140%
multipliers = [0.70, 0.80, 0.90, 1.00, 1.10, 1.20, 1.30, 1.40, 1.50]

ws4.cell(4, 1, "מכפיל עלות יחידה").font = bold(10, WHITE)
ws4.cell(4, 1).fill = fill(MID_BLUE); ws4.cell(4, 1).alignment = center(); ws4.cell(4, 1).border = border()
ws4.cell(4, 2, "עלות הון (₪)").font = bold(10, WHITE)
ws4.cell(4, 2).fill = fill(MID_BLUE); ws4.cell(4, 2).alignment = center(); ws4.cell(4, 2).border = border()
ws4.cell(4, 3, "עלות הון + רזרבה (₪)").font = bold(10, WHITE)
ws4.cell(4, 3).fill = fill(MID_BLUE); ws4.cell(4, 3).alignment = center(); ws4.cell(4, 3).border = border()
ws4.cell(4, 4, "תחזוקה שנתית (₪)").font = bold(10, WHITE)
ws4.cell(4, 4).fill = fill(MID_BLUE); ws4.cell(4, 4).alignment = center(); ws4.cell(4, 4).border = border()
ws4.cell(4, 5, "עלות 10 שנים (₪)").font = bold(10, WHITE)
ws4.cell(4, 5).fill = fill(MID_BLUE); ws4.cell(4, 5).alignment = center(); ws4.cell(4, 5).border = border()

# Compute base capital from actual data
base_cap = sum(default_costs.get(r, 400_000) for _, _, r in shelter_points)
maint_annual = 15_000 * total_shelters
contingency = 0.15

for i, mult in enumerate(multipliers):
    r = i + 5
    cap     = base_cap * mult
    cap_res = cap * (1 + contingency)
    total10 = cap_res + maint_annual * 10
    bg = "E8F5E9" if mult == 1.00 else ("FFF9C4" if mult < 1.00 else "FFEBEE")
    label = f"{int(mult*100)}%" + (" ← בסיס" if mult == 1.00 else "")
    ws4.cell(r, 1, label).font = bold(10) if mult == 1.00 else normal(10)
    ws4.cell(r, 1).fill = fill(bg); ws4.cell(r, 1).alignment = center(); ws4.cell(r, 1).border = border()
    for col, val in [(2, cap), (3, cap_res), (4, maint_annual), (5, total10)]:
        c = ws4.cell(r, col, round(val))
        c.number_format = '#,##0 ₪'
        c.font = bold(10) if mult == 1.00 else normal(10)
        c.fill = fill(bg); c.alignment = center(); c.border = border()

# Note
ws4.merge_cells(f"A{len(multipliers)+6}:F{len(multipliers)+6}")
c = ws4.cell(len(multipliers)+6, 1,
    "הערה: הטבלה מבוססת על עלויות בסיס ולא מקושרת לתאי הקלט בלוח הבקרה. לחישוב עם ערכים מותאמים — עדכן את לוח הבקרה.")
c.font = normal(9, "777777"); c.alignment = right()

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = BASE / "budget_shelters.xlsx"
wb.save(out_path)
print(f"\nנשמר: {out_path}")
print("גיליונות:")
print("  1. לוח בקרה   — פרמטרים עריכים + סיכום")
print("  2. מיגוניות    — רשימה מלאה")
print("  3. לפי אזור   — פירוט אזורי")
print("  4. רגישות      — טבלת what-if")
